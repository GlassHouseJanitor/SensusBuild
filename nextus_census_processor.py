#!/usr/bin/env python3
import os
import pandas as pd
import numpy as np
import glob
import re
from datetime import datetime
import calendar
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

def process_census_files(input_folder='uploads',
                         output_folder='uploads',
                         month=3, year=2025):
    """
    Process daily attendance CSV files and generate a formatted census report.
    """
    print(f"Processing census data for {calendar.month_name[month]} {year}...")
    
    # Create output folder if it doesn't exist
    os.makedirs(output_folder, exist_ok=True)
    
    # Define month details
    month_name = calendar.month_name[month]
    _, days_in_month = calendar.monthrange(year, month)
    
    # ========== STEP 1: FIND CSV FILES ==========
    print(f"Looking for CSV files in {input_folder}...")
    
    # Find all CSV files in the input folder
    csv_files = glob.glob(os.path.join(input_folder, "*.csv"))
    if not csv_files:
        print(f"No CSV files found in {input_folder}")
        return
    
    print(f"Found {len(csv_files)} CSV files:")
    for file in csv_files:
        print(f"  - {os.path.basename(file)}")
    
    # ========== STEP 2: PROCESS ALL CSV FILES ==========
    print("\nProcessing all CSV files...")
    
    # Dictionary to store patient data
    patient_data = {}
    
    # Program code mapping - map your program codes to the ones in the template
    program_map = {
        'SUD-PHP': 'PHP',
        'SUD-OP': 'OP',
        'MH-PHP': 'MHPHP',
        'MH-IOP': 'MHIOP',
        'PHP': 'PHP',
        'IOP': 'IOP',
        'OP': 'OP',
        'MHPHP': 'MHPHP',
        'MHIOP': 'MHIOP'
    }
    
    # Process each CSV file
    for file_path in csv_files:
        try:
            filename = os.path.basename(file_path)
            print(f"Processing {filename}")
            
            # Extract date from filename
            date_match = re.search(r'(\d{4}[-_]?\d{2}[-_]?\d{2})', filename)
            if not date_match:
                print(f"  Could not extract date from filename, skipping")
                continue
                
            date_str = date_match.group(1).replace('_', '-')
            try:
                file_date = datetime.strptime(date_str, '%Y-%m-%d')
                print(f"  Extracted date from filename: {file_date.strftime('%Y-%m-%d')}")
            except ValueError:
                print(f"  Invalid date format in filename, skipping")
                continue
            
            # Skip if not in target month/year
            if file_date.month != month or file_date.year != year:
                print(f"  File date not in target month/year, skipping")
                continue
            
            # Read the CSV file
            df = pd.read_csv(file_path)
            print(f"  File contains {len(df)} records")
            
            day_of_month = file_date.day
            
            # Process each row in the file
            records_processed = 0
            
            for _, row in df.iterrows():
                try:
                    # Skip rows without a Name
                    if 'Name' not in df.columns or pd.isna(row['Name']):
                        continue
                    
                    # Get patient name and split into first/last
                    full_name = row['Name'].strip()
                    name_parts = full_name.split()
                    
                    if len(name_parts) < 2:
                        continue  # Skip if can't parse name
                    
                    last_name = name_parts[-1]
                    first_name = ' '.join(name_parts[:-1])
                    
                    # Get patient MR number for ID
                    mr_number = str(row.get('MR', '')).strip()
                    
                    # Create unique ID
                    unique_id = mr_number if mr_number else f"{last_name}_{first_name}"
                    
                    # Get program and map to correct code
                    program = row.get('Program', '')
                    if isinstance(program, str):
                        program = program.strip()
                        mapped_program = program_map.get(program, program)
                    else:
                        mapped_program = ''
                    
                    # Get attendance status
                    status = row.get('Status', '')
                    
                    # Determine service code based on status and program
                    service_code = ''
                    if status == 'Present' and mapped_program:
                        service_code = mapped_program
                    elif status == 'Absent':
                        service_code = 'X'  # X represents No Programming
                    
                    if not service_code:
                        continue  # Skip if no service to record
                    
                    # Get payment method
                    payment = row.get('Payment Method', '')
                    
                    # Create patient record if it doesn't exist
                    if unique_id not in patient_data:
                        patient_data[unique_id] = {
                            'last_name': last_name,
                            'first_name': first_name,
                            'admit_date': row.get('Admission', ''),
                            'payer_source': payment,
                            'program': mapped_program,
                            'icd10': '',  # No ICD-10 in your data
                            'ur_review': f"{row.get('UR Loc', '')} - Next review: {row.get('Next Review', '')}",
                            'billing_comments': row.get('Comment', ''),
                            'services': {}
                        }
                    
                    # Add service for this day
                    patient_data[unique_id]['services'][day_of_month] = service_code
                    
                    records_processed += 1
                
                except Exception as e:
                    print(f"  Error processing row: {e}")
            
            print(f"  Processed {records_processed} records with data in {month_name} {year}")
                
        except Exception as e:
            print(f"Error processing {filename}: {e}")
    
    # Summarize what we found
    print(f"\nFound data for {len(patient_data)} patients in {month_name} {year}")
    
    # ========== STEP 3: CREATE EXCEL WORKBOOK ==========
    print("\nCreating Excel workbook...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month_name} {year} Census"
    
    # Define styles
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    light_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    header_font = Font(name='Calibri', size=11, bold=True)
    normal_font = Font(name='Calibri', size=10)
    blue_font = Font(name='Calibri', size=11, bold=True, color="0000FF")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    wrap_text = Alignment(wrapText=True, vertical='center')
    
    # ========== HEADER SECTION ==========
    # Company header
    ws.merge_cells('A1:J1')
    ws['A1'] = "Glass House Recovery LLC"
    ws['A1'].font = Font(name='Calibri', size=12, bold=True)
    ws['A1'].alignment = left_align
    
    # Address info
    ws.merge_cells('A2:J2')
    ws['A2'] = "8318 Forrest st STE 100"
    ws['A2'].font = normal_font
    ws['A2'].alignment = left_align
    
    ws.merge_cells('A3:J3')
    ws['A3'] = "Ellicott City, MD 21043-5148"
    ws['A3'].font = normal_font
    ws['A3'].alignment = left_align
    
    # Contact info
    ws.merge_cells('A4:J4')
    ws['A4'] = "Medical Director: Tetyana Evans CRNP"
    ws['A4'].font = normal_font
    ws['A4'].alignment = left_align
    
    ws.merge_cells('A5:J5')
    ws['A5'] = "Clinical Director: Allison Moberly, LCPC"
    ws['A5'].font = normal_font
    ws['A5'].alignment = left_align
    
    # Rest of the formatting code remains the same...
    
    # Define styles and create legend sections, month headers, etc.
    # (Using the existing code from my previous answer)
    
    # ========== LEGEND SECTION ==========
    # First legend section
    ws.merge_cells('L1:N1')
    ws['L1'] = "LEGEND"
    ws['L1'].font = header_font
    ws['L1'].alignment = center_align
    
    # Program legend table
    legend_data = [
        ("PHP", "Partial Hospitalization"),
        ("IOP", "Intensive Outpatient Program"),
        ("INT", "Intake"),
        ("MHPHP", "Partial Hospitalization primary mental health"),
        ("MHIOP", "Intensive Outpatient Program primary mental health"),
        ("MHINT", "Intake Primary Mental Health"),
        ("OP", "OP Group 45 min"),
        ("FUP15", "Medical Followup - 15min"),
        ("IPEM", "Initial Psych evaluation with medical services"),
        ("IND 45", "OP Individual 45 min"),
        ("FUP25", "Medical Followup - 25min"),
        ("x", "No Programming"),
        ("DIS", "Outpatient Discharge"),
        ("FUP45", "Medical Followup - 45min"),
        ("(Blank)", "No billable service")
    ]
    
    for idx, (code, desc) in enumerate(legend_data):
        row = idx + 2  # Start from row 2
        
        ws.cell(row=row, column=12).value = code  # Column L
        ws.cell(row=row, column=12).font = normal_font
        ws.cell(row=row, column=12).alignment = center_align
        ws.cell(row=row, column=12).border = thin_border
        
        ws.cell(row=row, column=13).value = desc  # Column M
        ws.cell(row=row, column=13).font = normal_font
        ws.cell(row=row, column=13).alignment = left_align
        ws.cell(row=row, column=13).border = thin_border
    
    # Add copyright notice
    ws.merge_cells('L17:N17')
    ws['L17'] = "Process and Template Proprietary Property of Nextus Billing Solutions 1-1-18"
    ws['L17'].font = Font(name='Calibri', size=8, italic=True)
    ws['L17'].alignment = center_align
    
    # Claims legend section
    ws.merge_cells('O1:S1')
    ws['O1'] = "CLAIMS LEGEND"
    ws['O1'].font = header_font
    ws['O1'].alignment = center_align
    
    # Claims legend data
    claims_data = [
        ("Paid", "PHP/UA", "", "Paid to Patient", "PHP/UA"),
        ("", "PHP/UA (see note) [1]", "Submitted", "", ""),
        ("Needs resubmission", "PHP/UA [2]", "Partial P2P", "PHP/UA", ""),
        ("Final Denial", "PHP/UA", "Partial Payment", "PHP/UA", "Unbillable Service"),
        ("Pending Submission", "PHP [4]", "Partial Final Denial", "PHP/CM", "Auth obtained")
    ]
    
    for idx, (status, code1, desc1, status2, code2) in enumerate(claims_data):
        row = idx + 2  # Start from row 2
        
        # First column - Status
        ws.cell(row=row, column=15).value = status  # Column O
        ws.cell(row=row, column=15).font = normal_font
        ws.cell(row=row, column=15).alignment = center_align
        ws.cell(row=row, column=15).border = thin_border
        
        # Second column - Code
        ws.cell(row=row, column=16).value = code1  # Column P
        ws.cell(row=row, column=16).font = normal_font
        ws.cell(row=row, column=16).alignment = center_align
        ws.cell(row=row, column=16).border = thin_border
        if "PHP/UA" in code1 or code1 == "PHP [4]":
            ws.cell(row=row, column=16).fill = yellow_fill
        
        # Third column - Description
        ws.cell(row=row, column=17).value = desc1  # Column Q
        ws.cell(row=row, column=17).font = normal_font
        ws.cell(row=row, column=17).alignment = left_align
        ws.cell(row=row, column=17).border = thin_border
        
        # Fourth column - Paid to Patient
        ws.cell(row=row, column=18).value = status2  # Column R
        ws.cell(row=row, column=18).font = normal_font
        ws.cell(row=row, column=18).alignment = center_align
        ws.cell(row=row, column=18).border = thin_border
        
        # Fifth column - Code for Paid to Patient
        ws.cell(row=row, column=19).value = code2  # Column S
        ws.cell(row=row, column=19).font = normal_font
        ws.cell(row=row, column=19).alignment = center_align
        ws.cell(row=row, column=19).border = thin_border
        if "PHP/UA" in code2:
            ws.cell(row=row, column=19).fill = yellow_fill
        elif code2 == "Auth obtained":
            ws.cell(row=row, column=19).fill = green_fill
        elif code2 == "Unbillable Service":
            ws.cell(row=row, column=19).fill = red_fill
    
    # ========== MONTH NAME AND DAY HEADERS ==========
    # Month/Year header
    ws.merge_cells('A6:G6')
    ws['A6'] = f"{month_name} {year}"
    ws['A6'].font = header_font
    ws['A6'].alignment = center_align
    ws['A6'].fill = light_blue_fill
    
    # Day of week headers
    day_labels = ["Sat", "Sun", "Mon", "Tue", "Wed", "Thu", "Fri"] * 5
    for day in range(1, days_in_month + 1):
        col = 7 + day
        cell = ws.cell(row=6, column=col)
        cell.value = day_labels[(day - 1) % 7]
        cell.font = normal_font
        cell.alignment = center_align
        cell.fill = light_blue_fill
        cell.border = thin_border
    
    # ========== MAIN TABLE HEADERS ==========
    header_row = 7
    headers = ["Last Name", "First Name", "Admit Date", "Payer Source",
               "Program", "ICD 10", "Fee"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = light_blue_fill
        cell.border = thin_border
    
    # Day number headers
    for day in range(1, days_in_month + 1):
        col = len(headers) + day
        cell = ws.cell(row=header_row, column=col)
        cell.value = day
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = light_blue_fill
        cell.border = thin_border
    
    # UR Comments and Billing Comments headers
    ur_col = len(headers) + days_in_month + 1
    billing_col = len(headers) + days_in_month + 2
    
    ws.cell(row=header_row, column=ur_col).value = "UR Comments"
    ws.cell(row=header_row, column=ur_col).font = header_font
    ws.cell(row=header_row, column=ur_col).alignment = center_align
    ws.cell(row=header_row, column=ur_col).fill = light_blue_fill
    ws.cell(row=header_row, column=ur_col).border = thin_border
    
    ws.cell(row=header_row, column=billing_col).value = "Billing Comments"
    ws.cell(row=header_row, column=billing_col).font = header_font
    ws.cell(row=header_row, column=billing_col).alignment = center_align
    ws.cell(row=header_row, column=billing_col).fill = light_blue_fill
    ws.cell(row=header_row, column=billing_col).border = thin_border
    
    # ========== POPULATE PATIENT DATA ==========
    print("Populating patient data...")
    
    # Separate patients by insurance type
    standard_patients = {}
    medicaid_patients = {}
    
    for patient_id, patient in patient_data.items():
        payer = patient.get('payer_source', '').lower()
        if 'medicaid' in payer:
            medicaid_patients[patient_id] = patient
        else:
            standard_patients[patient_id] = patient
    
    # Function to add a patient row
    def add_patient_row(patient, row_num):
        # Basic patient info
        ws.cell(row=row_num, column=1).value = patient.get('last_name', '')
        ws.cell(row=row_num, column=2).value = patient.get('first_name', '')
        ws.cell(row=row_num, column=3).value = patient.get('admit_date', '')
        ws.cell(row=row_num, column=4).value = patient.get('payer_source', '')
        ws.cell(row=row_num, column=5).value = patient.get('program', '')
        ws.cell(row=row_num, column=6).value = patient.get('icd10', '')
        
        # Format patient info cells
        for col in range(1, 8):
            cell = ws.cell(row=row_num, column=col)
            cell.font = normal_font
            cell.border = thin_border
            
            # Set alignment based on column
            if col == 6:  # ICD-10 column needs wrapping
                cell.alignment = wrap_text
            else:
                cell.alignment = left_align
        
        # Add service data for each day
        for day in range(1, days_in_month + 1):
            col = len(headers) + day
            cell = ws.cell(row=row_num, column=col)
            
            # Skip if the cell is a merged cell
            if isinstance(cell, MergedCell):
                continue
                
            # Get service for this day
            service = patient.get('services', {}).get(day, '')
            cell.value = service
            
            # Format the cell
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = center_align
            
            # Apply yellow highlight for service codes
            if service and service != 'X' and not service.startswith('X'):
                cell.fill = yellow_fill
        
        # Add UR Review and Billing Comments
        ur_cell = ws.cell(row=row_num, column=ur_col)
        ur_cell.value = patient.get('ur_review', '')
        ur_cell.font = normal_font
        ur_cell.border = thin_border
        ur_cell.alignment = wrap_text
        
        billing_cell = ws.cell(row=row_num, column=billing_col)
        billing_cell.value = patient.get('billing_comments', '')
        billing_cell.font = normal_font
        billing_cell.border = thin_border
        billing_cell.alignment = wrap_text
    
    # Add standard patients
    current_row = header_row + 1
    for patient_id, patient in sorted(standard_patients.items(), key=lambda x: x[1].get('last_name', '')):
        add_patient_row(patient, current_row)
        current_row += 1
    
    # Add "Medicaid Patients Below" section
    if medicaid_patients:
        # Add divider row
        medicaid_row = current_row
        ws.merge_cells(f'A{medicaid_row}:G{medicaid_row}')
        ws.cell(row=medicaid_row, column=1).value = "Medicaid Patients Below"
        ws.cell(row=medicaid_row, column=1).font = blue_font
        ws.cell(row=medicaid_row, column=1).alignment = center_align
        ws.cell(row=medicaid_row, column=1).fill = light_blue_fill
        
        # Add borders to merged cells
        for col in range(1, 8):
            ws.cell(row=medicaid_row, column=col).border = thin_border
        
        current_row += 1
        
        # Add Medicaid patients
        for patient_id, patient in sorted(medicaid_patients.items(), key=lambda x: x[1].get('last_name', '')):
            add_patient_row(patient, current_row)
            current_row += 1
    
    # ========== ADJUST COLUMN WIDTHS ==========
    column_widths = {
        'A': 15,  # Last Name
        'B': 15,  # First Name
        'C': 12,  # Admit Date
        'D': 20,  # Payer Source
        'E': 10,  # Program
        'F': 20,  # ICD-10
        'G': 8,   # Fee
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Set width for day columns
    for day in range(1, days_in_month + 1):
        col_letter = get_column_letter(7 + day)
        ws.column_dimensions[col_letter].width = 8
    
    # Set width for comment columns
    ws.column_dimensions[get_column_letter(ur_col)].width = 40
    ws.column_dimensions[get_column_letter(billing_col)].width = 25
    
    # ========== FOOTNOTES WORKSHEET ==========
    footnotes_ws = wb.create_sheet(title="Footnotes")
    
    footnotes = [
        "[1] 1/31/99 - Status note",
        "[2] 1/31/99 - Status note",
        "[3] 1/31/99 - Status note",
        "[4] 1/31/99 - Status note",
        "[5] 3/11/25- Pending auth- GN",
        "[6] 11/18/24- See billing instruction on SCA. - CC",
        "[7] 3/4/25- PENDING LOC CLARIFICATION-GN",
        "[8] 3/4/25- PENDING LOC CLARIFICATION-GN",
        "[9] 3/4/25- PENDING LOC CLARIFICATION-GN",
        "[10] 3/4/25- PENDING LOC CLARIFICATION-GN",
        "[11] 3/4/25- PENDING LOC CLARIFICATION-GN",
        "[12] 3/4/25- PENDING LOC CLARIFICATION-GN",
        "[13] 3/4/25- PENDING LOC CLARIFICATION-GN",
        "[14] 3/11/25- PENDING AUTH- GN",
        "[15] 3/11/25- PENDING AUTH- GN",
        "[16] 3/11/25- PENDING AUTH- GN"
    ]
    
    for i, note in enumerate(footnotes, 1):
        footnotes_ws.cell(row=i, column=1).value = note
        footnotes_ws.cell(row=i, column=1).font = normal_font
    
    # ========== SAVE WORKBOOK ==========
    output_path = os.path.join(output_folder, f"Census_{month_name}_{year}.xlsx")
    wb.save(output_path)
    print(f"Census report saved to {output_path}")
    
    return output_path

if __name__ == "__main__":
    # Process files for March 2025 by default
    process_census_files(month=3, year=2025)
